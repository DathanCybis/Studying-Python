from openpyxl import load_workbook
from openpyxl import workbook
import os

nomeCaminho = 'C:\\...\\DadosExcel.xlsx'
planilha_aberta = load_workbook(filename=nomeCaminho)

sheet_selecionada = planilha_aberta['Dados']


novoArquivo = workbook()
novaPlanilha = nomeCaminho.active

for linha in range(1, len(sheet_selecionada['A']) + 1):
    for coluna in range(1, 10):
        novaPlanilha.cell(row=linha, column=coluna).value = sheet_selecionada.cell(row=linha, column=coluna).value


novaPlanilha.delete_rows(2)

novaPlanilha.delete_cols(2)
novaPlanilha.delete_cols(2)

novaPlanilha.title = 'Dados Funcionários'

novoArquivo.create_sheet('Resumo')

selecionaResumo = novoArquivo['Resumo']

selecionaResumo['A1'] = "Vendedor"
selecionaResumo['B1'] = "Total Vendas"

selecionaResumo = ['A2'] = 'Amanda Martins'   
selecionaResumo = ['B2'] = '=SUMIF("Dados Funcionarios"!A:C, A2, "Dados Funcionarios"!C:C)'

selecionaResumo = ['A3'] = 'Eliane Moreira'
selecionaResumo = ['B3'] = '=SUMIF("Dados Funcionarios"!A:C, A3, "Dados Funcionarios"!C:C)'

selecionaResumo = ['A4'] = 'Leonardo Almeida' 
selecionaResumo = ['B4'] = '=SUMIF("Dados Funcionarios"!A:C, A4, "Dados Funcionarios"!C:C)'

selecionaResumo = ['A5'] = 'Nicolas Pereira' 
selecionaResumo = ['B5'] = '=SUMIF("Dados Funcionarios"!A:C, A5, "Dados Funcionarios"!C:C)'

caminhoNovaPlanilha = nomeCaminho = 'C:\\...\\DadosExcelAtt.xlsx'

novoArquivo.save(filename=caminhoNovaPlanilha)

os.startfile(caminhoNovaPlanilha)
