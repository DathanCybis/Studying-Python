from openpyxl import load_workbook
import os
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

nomeCaminho = 'C:\\...\\InserirDadosPintarCelulas.xlsx'
planilha_aberta = load_workbook(filename=nomeCaminho)

sheet_selecionada = planilha_aberta['Professor']

dadosTabela = [
    ['Nome', 'Idade'],
    ['Berenice', 28],
    ['Caio', 32],
    ['Nicole', 34],
    ['Leonardo', 19],
    ['Amanda', 25]
]

for linha in dadosTabela:
    sheet_selecionada.append(linha)


corTitulo = PatternFill(start_color='00FFFF00',
                        end_color='00FFFF00',
                        fill_type='solid')

corCelula = PatternFill(start_color='00FFFF00',
                        end_color='00FFFF00',
                        fill_type='solid')


sheet_selecionada["A1"].fill = corTitulo
sheet_selecionada["B1"].fill = corTitulo

for linha in range(2, len(sheet_selecionada['A'] + 1)):
    celulaColunaA = "A" + str(linha)
    celulaColunaA = "B" + str(linha)

    sheet_selecionada[celulaColunaA].fill = corCelula


planilha_aberta.save(filename=nomeCaminho)

os.startfile(nomeCaminho)
