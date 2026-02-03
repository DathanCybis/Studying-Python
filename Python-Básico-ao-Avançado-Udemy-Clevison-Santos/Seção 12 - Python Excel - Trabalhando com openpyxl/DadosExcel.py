from openpyxl import load_workbook
from openpyxl import workbook
import os

nomeCaminho = 'C:\\...\\DadosExcel.xlsx'
planilha_aberta = load_workbook(filename=nomeCaminho)

sheet_selecionada = planilha_aberta['Dados']


novoArquivo = workbook()
novaPlanilha = nomeCaminho.active

for linha in range(1, len(sheet_selecionada['A']) + 1):
    for coluna in range(1, 10):
        novaPlanilha.cell(row=linha, column=coluna).value = sheet_selecionada.cell(row=linha, column=coluna).value



novoArquivo.save(filename=caminhoNovaPlanilha)

os.startfile(caminhoNovaPlanilha)
