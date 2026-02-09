from openpyxl import load_workbook
from openpyxl import Workbook
import os

nomeArquivo = "C:\\...\\...\\DividindoArquivos.xlsx"

planilha_aberta = load_workbook(filename=nomeArquivo)

sheet_selecionada = planilha_aberta['Dados']

criandoNovoExcel = Workbook()

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

for linha in range(2, totalLinha):
    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:
        linhaSheet = len(selecionaSheet['A'] + 1)
        celulaColunaA = "A" + str(linhaSheet)
        celulaColunaB = "B" + str(linhaSheet)
        celulaColunaC = "C" + str(linhaSheet)

        selecionaSheet[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheet[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheet[celulaColunaC] = sheet_selecionada['C%s' % linha].value

    else:
        sheet_resumo = planilha_aberta.create_sheet(title=nomeAtual)

        selecionaSheet = planilha_aberta[nomeAtual]

        nomeAtual = sheet_selecionada['A%s' % linha].value

        nova_planilha = criandoNovoExcel.active

        nova_planilha.title = "Vendas"

        caminhoNovaPlanilha = "C:\\...\\...\\DividindoArquivos.xlsx"

        selecionaSheet = criandoNovoExcel['Vendas']

        selecionaSheet['A1'] = "Vendedor"
        selecionaSheet['B1'] = "Produtos"
        selecionaSheet['C1'] = "Vendas"

        selecionaSheet['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheet['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheet['C2'] = sheet_selecionada['C%s' % linha].value

        selecionaSheet.delete_rows(3, 100)

        criandoNovoExcel.save(filename=caminhoNovaPlanilha)

planilha_aberta.save(filename=nomeArquivo)

os.startfile(nomeArquivo)
