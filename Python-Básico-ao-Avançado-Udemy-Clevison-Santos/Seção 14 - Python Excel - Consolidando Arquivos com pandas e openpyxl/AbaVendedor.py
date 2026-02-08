from openpyxl import load_workbook
import os

nomeArquivo = "C:\\...\\...\\Quebrar.xlsx"

planilha_aberta = load_workbook(filename=nomeArquivo)

sheet_selecionada = planilha_aberta['Dados']

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

for linha in range(2, totalLinha):
    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:
        linhaSheet = len(sheet_selecionada2['A'] + 1)
        celulaColunaA = "A" + str(linhaSheet)
        celulaColunaB = "B" + str(linhaSheet)
        celulaColunaC = "C" + str(linhaSheet)

    else:
        sheet_resumo = planilha_aberta.create_sheet(title=nomeAtual)

        sheet_selecionada2 = planilha_aberta[nomeAtual]

        nomeAtual = sheet_selecionada['A%s' % linha].value

        sheet_selecionada2['A1'] = "Vendedor"
        sheet_selecionada2['B1'] = "Produtos"
        sheet_selecionada2['C1'] = "Vendas"

        sheet_selecionada2['A2'] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2['B2'] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2['B2'] = sheet_selecionada['C%s' % linha].value

planilha_aberta.save(filename=nomeArquivo)

os.startfile(nomeArquivo)
